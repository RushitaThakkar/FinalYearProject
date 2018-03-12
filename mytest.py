
import csv
import xml.etree.ElementTree as ET
import re
import networkx as nx
import matplotlib.pyplot as plt
from collections import namedtuple
from bs4 import BeautifulSoup
from time import sleep
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import csv
import openpyxl
from openpyxl import load_workbook
wbnew = openpyxl.Workbook()
wbnew.save('new_excel.xlsx')
ws2 = wbnew.active
mydata = set()
mycount = 0
def parseXML(xmlfile):

    initialize_fields()
    # create element tree object
    tree = ET.parse(xmlfile)

    # get root element
    root = tree.getroot()

    # create empty list for news items
    data = []
    outgoingUrl = []
    max = ws2.max_row + 1
    for item in root.findall('./Request'):
        # empty news dictionary
        request = {}
        ownPath = ""
        URL,host, host1,contentType = "", "","",""
        statusCode=100
        # iterate child elements of item
        for child in item:
            request[child.tag] = child.text
            if child.tag == "Hostname":
                host = child.text
                host = host.strip()
                host1 = "http://"+host+":80"
            if child.tag == "Url":
                URL = child.text
                req1 = requests.get(URL, allow_redirects=False)
                sessions = requests.Session()
                retry = Retry(connect=3, backoff_factor=0.5)
                adapter = HTTPAdapter(max_retries=retry)
                r = sessions.get(URL)
                requestMethod = r.request.method
                server = r.headers['Server']
                contentType = r.headers['Content-Type']
                contentLength = r.headers['content-length']
                URL = URL.strip()
                URL = URL[len(host1):]
                URL = check_both(URL)
            if child.tag == "RequestHeader":
                referer = processRequestHeaderReferer(child.text,host)
            if child.tag == "ResponseHeader":
                child.text = child.text.strip()
                statusCode = child.text[9:12]
                print(statusCode)
            if child.tag == "ResponseData":
                if not child.text:
                    continue
                soup = BeautifulSoup(child.text, 'html.parser')
                for link in soup.find_all('a'):
                    if link.get('href'):
                        a=link.get('href')
                        a = check_both(a)
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,0,contentType,contentLength)
                        max += 1
                for link in soup.find_all('script'):
                    if link.get('src'):
                        a=link.get('src')
                        a = check_both(a)
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,0,contentType,contentLength)
                        max += 1
                for link in soup.find_all('form'):
                    if link.get('action'):
                        a=link.get('action')
                        a = check_both(a)
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,1,contentType,contentLength)
                        max += 1
                for link in soup.find_all('meta'):
                    if link.get('URL'):
                        a=link.get('URL')
                        a = check_both(a)
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,0,contentType,contentLength)
                        max += 1
            # special checking for namespace object content:media

        # append news dictionary to news items list

def processRequestHeaderReferer(child,hostname):
    referer = re.findall(r'Referer: http://', child)
    if len(referer)==0:
        return ""
    referer = referer[0]+ hostname
    refererString = (re.findall(r'Referer: http://.*Accept-En',child))
    refererString = refererString[0]
    return refererString[len(referer):-9]

def initialize_fields():
    ws2.cell(row=1, column=1).value = "Source URL"
    ws2.cell(row=1, column=2).value = "Outgoing URL"
    ws2.cell(row=1, column=3).value = "Request Method"
    ws2.cell(row=1, column=4).value = "Host Name"
    ws2.cell(row=1, column=5).value = "Status Code"
    ws2.cell(row=1, column=6).value = "Server"
    ws2.cell(row=1, column=7).value = "Referer"
    ws2.cell(row=1, column=8).value = "Form"
    ws2.cell(row=1, column=9).value = "Content-type"
    ws2.cell(row=1, column=10).value = "Content-length"
def append_new(max,URL,a,requestMethod,Host,statusCode,server, referer,form,ct,cl):
    mydata.add((URL,a,requestMethod,Host,statusCode,server, referer,form,ct,cl))
    j=0

def check_file_type(a):
    print(type(a))
    if ".css" in a or ".png" in a or ".ico" in a or ".woff" in a or ".woff2" in a or ".gif" in a or ".txt" in a:
        return False
    else:
        return True
def check_path(a):
    if a[0]=='/':
        return a
    else:
        return "/"+a
def check_suffix(a):
    if a[-1]=='/':
        return a
    else:
        return a+"/"
def check_both(a):
    a = check_path(a)
    a = check_suffix(a)
    return a
def main():

    # parse xml file
    parseXML('test123.xml')
    print(len(mydata))
    outer,inner=1,0
    for i in mydata:
        outer+=1
        for j in i:
            inner+=1
            if inner==11:
                inner=1
            ws2.cell(row=outer, column=inner).value = j
    wbnew.save('new_excel.xlsx')

if __name__ == "__main__":
    # calling main function
    main()